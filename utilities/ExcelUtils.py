import openpyxl

def getRowCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_row)

def getColCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_column)

def readData(file,SheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(row=rowNum, column=colNum).value

def writeData(file,SheetName,rowNum,colNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(row=rowNum, col=colNum).value = data
    workbook.save(file)
